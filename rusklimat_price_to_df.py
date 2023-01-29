import pandas as pd
import openpyxl

# df = pd.read_excel('data/price_list.xlsx')
FILE = 'data/price_list_test.xlsx'
WORKSHEET = 'Worksheet'


def parce_sheet(ws):
    levels = []
    level_items = []

    for row in range(ws.min_row, ws.max_row + 1):
        current_level = int(ws.row_dimensions[row].outline_level)
        next_level = int(ws.row_dimensions[row+1].outline_level)

        def textconv(cell):
            text = cell.value.strip() if cell.value else ''
            return text

        def priceconv(price):
            try:
                price_value = str(price.value.replace(' ', ''))
                price = float(price_value.replace(',', '.')) if price_value else 0.0
                return price
            except Exception as ex:
                return ''

        if current_level != 0:
            if next_level > current_level:
                levels.append(ws.cell(row, 1).value.strip())
            elif next_level == current_level:
                append_string = [len(levels)] + \
                                levels + \
                                [textconv(ws.cell(row, 1)), textconv(ws.cell(row, 5)), priceconv(ws.cell(row, 8))]
                # print(row, append_string)
                level_items.append(append_string)
            elif next_level < current_level:
                append_string = [len(levels)] + \
                                levels + \
                                [textconv(ws.cell(row, 1)), textconv(ws.cell(row, 5)), priceconv(ws.cell(row, 8))]
                level_items.append(append_string)
                for i in range(current_level-next_level):
                    if levels:
                        levels.pop()
    return level_items


def create_structure(lst: list) -> pd.DataFrame:
    max_level = max([x[0] for x in lst])
    for n, i in enumerate(lst):
        if i[0] < max_level:
            lst[n] = i[:i[0]+1] + ['']*(max_level-i[0]) + i[i[0]+1:]

    df = pd.DataFrame(lst, columns=['level', 'header1', 'header2', 'header3', 'id', 'name', 'price'])
    return df


def main():
    wb = openpyxl.open(FILE)
    lst = parce_sheet(wb[WORKSHEET])
    # print(lst)
    df = create_structure(lst)

    df.to_excel('price_to_df.xlsx', index=False)


if __name__ == '__main__':
    main()
